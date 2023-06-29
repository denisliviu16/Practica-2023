import requests
import openpyxl
import pandas as pd
from openpyxl import Workbook
from bs4 import BeautifulSoup
import re

url = 'https://en.wikipedia.org/wiki/List_of_sovereign_states_in_Europe_by_GDP_(nominal)'
response = requests.get(url)
content = response.text

excel_file_path = 'gdp.xlsx'
workbook = Workbook()
worksheet = workbook.active

soup = BeautifulSoup(content, 'lxml')
table = soup.find('table', {'class': 'wikitable'})

rows = table.find_all('tr')
for row_index, row in enumerate(rows):
    cells = row.find_all(['th', 'td'])
    
    for column_index, cell in enumerate(cells):
        value = cell.text.strip()
        value = re.sub(r'\[.*?\]', '', value) 
        if column_index == 1:
            value = re.sub('[^a-zA-Z ]', '', value)
        value = value.replace(",", "")
        worksheet.cell(row=row_index+1, column=column_index+1, value=value)

worksheet.cell(row=1, column=1, value="Rank")
workbook.save(excel_file_path)
df = pd.read_excel(excel_file_path)

df_sorted = df.sort_values(by="Country")

df_sorted.to_excel(excel_file_path, index=False)