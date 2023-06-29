import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://www.ef.com/wwen/blog/language/unbelievable-facts-make-you-seem-cultured/"
response = requests.get(url)
content = response.content

excel_file_path = 'facts.xlsx'
workbook = Workbook()
worksheet = workbook.active

soup = BeautifulSoup(content, 'lxml')
paragraphs = soup.find_all("p")

paragraphs = paragraphs[20:]

count=0
row_index=0
column_index=-1

for paragraph in paragraphs:
    
    if count % 10 == 0:
        column_index+=1
        row_index=0
    
    row_index+=1
    count+=1
    
    value = f"{row_index}. {paragraph.get_text().strip()}"
    worksheet.cell(row=row_index+1, column=column_index+1, value=value)
    
    
    if count == 50:
        break
    
worksheet.cell(row=1, column=1, value="Category 1: Nature")
worksheet.cell(row=1, column=2, value="Category 2: History")
worksheet.cell(row=1, column=3, value="Category 3: Art&Culture")
worksheet.cell(row=1, column=4, value="Category 4: People&Countries")    
worksheet.cell(row=1, column=5, value="Category 5: No way!Really?")        
workbook.save(excel_file_path)