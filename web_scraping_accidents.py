import requests
import pandas as pd
from openpyxl import Workbook
import openpyxl
from bs4 import BeautifulSoup

url = 'https://transport.ec.europa.eu/news-events/news/preliminary-2021-eu-road-safety-statistics-2022-03-28_en'
response = requests.get(url)
content = response.text

tables = pd.read_html(url)

table = tables[0] 

pd.set_option('display.max_columns', None)

table = table.iloc[:, 0:4]
table = table.iloc[:, 0:].apply(lambda x: x[1:])
table = table.drop(30)

excel_file_path = 'accidents.xlsx'
workbook = Workbook()

worksheet = workbook.active


for column_index, column_name in enumerate(table.columns):
    column_data = table[column_name]
    for row_index, value in enumerate(column_data):
        worksheet.cell(row=row_index+1, column=column_index+1, value=value)

worksheet.cell(row=1, column=1, value="Country")
workbook.save(excel_file_path)