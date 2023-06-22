import requests
import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup

url = 'https://en.m.wikipedia.org/wiki/Budget_of_the_European_Union'
response = requests.get(url)
content = response.text

soup = BeautifulSoup(content, 'lxml')
table = soup.find_all('table', {'class': 'wikitable'})
table = table[2]

excel_file_path = 'budget.xlsx'
workbook = Workbook()
worksheet = workbook.active

rows = table.find_all('tr')

for row_index, row in enumerate(rows):
    cells = row.find_all('td')
    
    for column_index, cell in enumerate(cells):
        value = cell.text.strip()
        worksheet.cell(row=row_index+1, column=column_index+1, value=value)

worksheet.cell(row=1, column=1, value="Member state")
worksheet.cell(row=1, column=2, value="Contribution")
workbook.save(excel_file_path)
