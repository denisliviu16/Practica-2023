import requests
import openpyxl
import pandas as pd
from openpyxl import Workbook
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook

url = 'https://en.wikipedia.org/wiki/List_of_countries_by_median_age'
response = requests.get(url)
content = response.text

excel_file_path = 'africa.xlsx'
workbook = Workbook()
worksheet = workbook.active

soup = BeautifulSoup(content, 'lxml')
table = soup.find('table', {'class': 'wikitable'})
rows = table.find_all('tr')

for row_index, row in enumerate(rows):
    cells = row.find_all(['th','td'])
    selected_columns = [0, 3, 4, 5]
    
    for column_index, cell in enumerate(cells):
        if column_index in selected_columns:
            value = cell.text.strip()
            worksheet.cell(row=row_index+1, column=column_index+1, value=value)

worksheet.delete_rows(1,2)

worksheet.cell(row=1, column=1, value="Country")
worksheet.cell(row=1, column=4, value="Combined")
worksheet.cell(row=1, column=5, value="Male")
worksheet.cell(row=1, column=6, value="Female")

workbook.save(excel_file_path)