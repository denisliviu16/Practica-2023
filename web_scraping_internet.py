import requests
import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup

url = 'https://en.wikipedia.org/wiki/List_of_countries_by_number_of_Internet_users'
response = requests.get(url)
content = response.text

soup = BeautifulSoup(content, 'lxml')
table = soup.find_all('table', {'class': 'wikitable'})
table = table[2]
rows = table.find_all('tr')

excel_file_path = 'internet.xlsx'
workbook = Workbook()
worksheet = workbook.active

for row_index, row in enumerate(rows):
    cells = row.find_all('td')
    cells = cells[:6]
    for column_index, cell in enumerate(cells):
        value = cell.text.strip()
        worksheet.cell(row=row_index+1, column=column_index+1, value=value)
        
worksheet.cell(row=1, column=1, value="Country")
worksheet.cell(row=1, column=2, value="Subregion")
worksheet.cell(row=1, column=3, value="Region")
worksheet.cell(row=1, column=4, value="Internet users")
worksheet.cell(row=1, column=5, value="Pct")
worksheet.cell(row=1, column=6, value="Population(2021)")
workbook.save(excel_file_path)